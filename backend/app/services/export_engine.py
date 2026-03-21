"""
导出引擎 - 发改委 XML + 商务部 Excel 模板映射生成
"""

import os
import io
import zipfile
from typing import Optional, Dict, List

from app.utils import utc_now

from lxml import etree
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.project import ProjectInvestment
from app.models.entity import EntityDomestic, EntityOverseas
from app.models.export_template import ExportTemplate, TemplateType


# 默认映射（当 DB 为空时使用）
DEFAULT_NDRC_FIELD_MAP = {
    "ProjectName": "project_name",
    "InvestorName": "company_name",
    "CreditCode": "uscc",
    "TargetCountry": "target_country",
    "OverseasCompanyNameCN": "overseas_name_cn",
    "OverseasCompanyNameEN": "overseas_name_en",
    "InvestmentAmount": "investment_amount",
    "Currency": "currency",
    "InvestmentPath": "investment_path",
    "IndustryCode": "industry_code",
    "OverseasIndustryCode": "overseas_industry_code",
    "NetAssets": "net_assets",
    "NetProfit": "net_profit",
}

DEFAULT_MOFCOM_COLUMNS = [
    ("项目名称", "project_name"),
    ("境内投资主体", "company_name"),
    ("统一社会信用代码", "uscc"),
    ("投资目的国/地区", "target_country"),
    ("境外企业名称(中文)", "overseas_name_cn"),
    ("境外企业名称(英文)", "overseas_name_en"),
    ("拟投资总额", "investment_amount"),
    ("币种", "currency"),
    ("投资路径", "investment_path"),
    ("境内行业代码", "industry_code"),
    ("境外行业代码", "overseas_industry_code"),
    ("注册资本", "registered_capital"),
    ("最近一年净资产", "net_assets"),
    ("最近一年净利润", "net_profit"),
    ("投资必要性说明", "purpose_description"),
]


class ExportEngine:
    """
    模板映射导出引擎
    - 发改委: XML 格式
    - 商务部: Excel 格式
    - 支持一键打包下载
    """

    @classmethod
    async def _get_ndrc_map(cls, db: AsyncSession) -> Dict[str, str]:
        """从 DB 获取 NDRC 字段映射，若为空则使用默认"""
        result = await db.execute(
            select(ExportTemplate)
            .where(ExportTemplate.template_type == TemplateType.NDRC)
            .where(ExportTemplate.is_active == True)
            .order_by(ExportTemplate.column_index)
        )
        templates = result.scalars().all()
        if not templates:
            return DEFAULT_NDRC_FIELD_MAP
        return {t.xml_tag: t.data_key for t in templates if t.xml_tag}

    @classmethod
    async def _get_mofcom_columns(cls, db: AsyncSession) -> List[tuple]:
        """从 DB 获取 MOFCOM 列配置，若为空则使用默认"""
        result = await db.execute(
            select(ExportTemplate)
            .where(ExportTemplate.template_type == TemplateType.MOFCOM)
            .where(ExportTemplate.is_active == True)
            .order_by(ExportTemplate.column_index)
        )
        templates = result.scalars().all()
        if not templates:
            return DEFAULT_MOFCOM_COLUMNS
        return [(t.display_name, t.data_key) for t in templates]

    @staticmethod
    async def _collect_data(db: AsyncSession, project: ProjectInvestment) -> dict:
        """收集并合并 Master Data"""
        data = {
            "project_name": project.project_name,
            "investment_amount": str(project.investment_amount or ""),
            "currency": project.currency
            if isinstance(project.currency, str)
            else project.currency.value
            if project.currency
            else "",
            "investment_path": project.investment_path
            if isinstance(project.investment_path, str)
            else project.investment_path.value
            if project.investment_path
            else "",
            "purpose_description": project.purpose_description or "",
        }

        if project.domestic_entity_id:
            domestic = await db.get(EntityDomestic, project.domestic_entity_id)
            if domestic:
                data.update(
                    {
                        "company_name": domestic.company_name,
                        "uscc": domestic.uscc,
                        "industry_code": domestic.industry_code or "",
                        "net_assets": str(domestic.net_assets or ""),
                        "net_profit": str(domestic.net_profit or ""),
                    }
                )

        if project.overseas_entity_id:
            overseas = await db.get(EntityOverseas, project.overseas_entity_id)
            if overseas:
                data.update(
                    {
                        "target_country": overseas.target_country,
                        "overseas_name_cn": overseas.overseas_name_cn,
                        "overseas_name_en": overseas.overseas_name_en or "",
                        "overseas_industry_code": overseas.overseas_industry_code or "",
                        "registered_capital": str(overseas.registered_capital or ""),
                    }
                )

        return data

    @classmethod
    async def generate_ndrc_xml(
        cls, db: AsyncSession, project: ProjectInvestment
    ) -> bytes:
        """生成发改委备案 XML"""
        data = await cls._collect_data(db, project)
        ndrc_map = await cls._get_ndrc_map(db)

        root = etree.Element("ODIProject")
        root.set("version", "1.0")
        root.set("timestamp", utc_now().isoformat())

        for xml_tag, data_key in ndrc_map.items():
            elem = etree.SubElement(root, xml_tag)
            elem.text = str(data.get(data_key, ""))

        # 资金来源
        if project.funding_source:
            funding_elem = etree.SubElement(root, "FundingSource")
            if isinstance(project.funding_source, dict):
                for key, value in project.funding_source.items():
                    sub_elem = etree.SubElement(funding_elem, key)
                    sub_elem.text = str(value)

        return etree.tostring(
            root, pretty_print=True, xml_declaration=True, encoding="UTF-8"
        )

    @classmethod
    async def generate_mofcom_excel(
        cls, db: AsyncSession, project: ProjectInvestment
    ) -> bytes:
        """生成商务部备案 Excel"""
        data = await cls._collect_data(db, project)
        mofcom_cols = await cls._get_mofcom_columns(db)

        wb = Workbook()
        ws = wb.active
        ws.title = "商务部备案信息"

        # 表头
        for col_idx, (header, _) in enumerate(mofcom_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)

        # 数据行
        for col_idx, (_, data_key) in enumerate(mofcom_cols, 1):
            ws.cell(row=2, column=col_idx, value=data.get(data_key, ""))

        # 自动调整列宽
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @classmethod
    async def generate_package(
        cls, db: AsyncSession, project: ProjectInvestment, export_type: str = "all"
    ) -> bytes:
        """一键打包下载"""
        timestamp = utc_now().strftime("%Y%m%d_%H%M%S")
        buffer = io.BytesIO()

        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            if export_type in ("all", "ndrc"):
                xml_bytes = await cls.generate_ndrc_xml(db, project)
                zf.writestr(f"发改委备案_{timestamp}.xml", xml_bytes)

            if export_type in ("all", "mofcom"):
                excel_bytes = await cls.generate_mofcom_excel(db, project)
                zf.writestr(f"商务部备案_{timestamp}.xlsx", excel_bytes)

        return buffer.getvalue()
